from pathlib import Path
from typing import Any, Callable, Generator, Optional

from zhtools.exceptions import ModuleRequired
from .interfaces import ReaderInterface

try:
    import openpyxl
except ImportError:
    raise ModuleRequired('openpyxl', '3.0.6')


class XlsxReader(ReaderInterface):

    def __init__(self, path: Path, sheet_name: str = None, sheet_idx: int = None):
        super().__init__(path)
        self.sheet_name = sheet_name
        self.sheet_idx = sheet_idx
        self._wb = None
        self.read_only = True

    @property
    def wb(self):
        if not self._wb:
            self._wb = openpyxl.open(self.path, read_only=self.read_only)

        if self._wb.read_only != self.read_only:
            self._wb = openpyxl.open(self.path, read_only=self.read_only)

        return self._wb

    def get_sheet_name(self) -> Optional[str]:
        if self.sheet_name:
            return self.sheet_name

        if not self.sheet_idx:
            return

        names = self.wb.get_sheet_names()
        return names[self.sheet_idx]

    def get_sheet(self):
        sheet_name = self.get_sheet_name()
        if sheet_name:
            sheet = self.wb.get_sheet_by_name(sheet_name)
        else:
            sheet = self.wb.active
        return sheet

    def close(self):
        self.wb.close()

    def read(self) -> list[list[Any]]:
        sheet = self.get_sheet()
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    def readline(self) -> Generator[tuple, None, None]:
        sheet = self.get_sheet()
        return sheet.iter_rows(values_only=True)

    def readlines(self) -> list[Any]:
        return list(self.readline())

    def read_col(self,
                 idx: int = 0,
                 filter: Callable[[Any], bool] = None
                 ) -> Generator[Any, None, None]:
        # ReadOnlyWorksheet is not supported iter_cols method.
        self.read_only = False
        sheet = self.get_sheet()
        for col in sheet.iter_cols(min_col=idx + 1,
                                   max_col=idx + 1,
                                   values_only=True):
            for val in col:
                if filter and not filter(val):
                    continue

                yield val
